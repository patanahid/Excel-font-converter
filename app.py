from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import requests
import json
import shutil


app = Flask(__name__)
def convert_kruti_to_unicode(input_text):

  url = 'https://hindi-font-converter-eight.vercel.app/api/unicode-krutidev'
  headers = {'Content-Type': 'application/json'}

  data = {
      'text': input_text,
      'format': 'json',
      'to_font': 'unicode'
  }

  response = requests.post(url, headers=headers, data=json.dumps(data))

  if response.status_code == 200:
    response_dict = json.loads(json.dumps(response.json()))

    # Extracting output_text
    output_text = response_dict.get("data", {}).get("output_text", "")
    print(output_text)
    return output_text
  else:
      return input_text

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_file():
    # Assuming your HTML file input has the name 'file'
    uploaded_file = request.files['file']
    os.makedirs("static")
    # Save the uploaded file
    uploaded_file.save("static/input.xlsx")

    # Indicate processing to the user
    os.rename("static/input.xlsx", "static/input_in_progress.xlsx")

    wb = load_workbook(filename="static/input_in_progress.xlsx")

    # Rest of your code for processing the Excel file
    sheets_to_duplicate = []

    Original_sheets = wb.sheetnames 

    for sheet_name in wb.sheetnames:
        sheets_to_duplicate.append(sheet_name)

    for sheet_name in sheets_to_duplicate:
        source = wb[sheet_name]
        target = wb.copy_worksheet(source)

        for row in target.iter_rows():
            for cell in row:
                if "devlys".upper() in cell.font.name.upper() or "kruti".upper() in cell.font.name.upper():
                    if cell.value is not None:
                        cell.value = convert_kruti_to_unicode(str(cell.value))
                        cell.font = Font(name="Calibri", size=cell.font.size, bold=cell.font.bold,
                                        italic=cell.font.italic, strikethrough=cell.font.strikethrough,
                                        underline=cell.font.underline, strike=cell.font.strike,
                                        color=cell.font.color, vertAlign=cell.font.vertAlign)

    for sheet in Original_sheets:
        del wb[sheet]

    # Save the modified workbook as output
    output_filename = "static/" + os.path.splitext(uploaded_file.filename)[0] + "-converted.xlsx"
    wb.save(output_filename)

    # Remove the in-progress indicator
    os.rename("static/input_in_progress.xlsx", "static/input.xlsx")

    try:
        return send_file(output_filename,
                        as_attachment=True,
                        download_name="output.xlsx")
    finally:
        global static_folder_path
        static_folder_path = os.path.join(os.getcwd(), "static")
        print(static_folder_path)
        def remove_files_in_static_folder():
            static_folder_path = os.path.join(os.getcwd(), "static")

        # Check if the folder exists
        if os.path.exists(static_folder_path) and os.path.isdir(static_folder_path):
            # List all files in the folder
            files_in_static = os.listdir(static_folder_path)

            # Iterate through the files and remove them
            for file_name in files_in_static:
                file_path = os.path.join(static_folder_path, file_name)
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        print(f"Removed: {file_path}")
                except Exception as e:
                    print(f"Error while removing {file_path}: {e}")

                print("All files in the 'static' folder have been removed.")
            else:
                print(
                    "The 'static' folder does not exist in the current working directory."
                )

        # Run the function
        remove_files_in_static_folder()
        shutil.rmtree(os.path.join(os.getcwd(), "static"))

if __name__ == '__main__':
    app.run(debug=False, host="0.0.0.0")
    def remove_files_in_static_folder():
        static_folder_path = os.path.join(os.getcwd(), "static")

# Check if the folder exists
if os.path.exists(static_folder_path) and os.path.isdir(static_folder_path):
    # List all files in the folder
    files_in_static = os.listdir(static_folder_path)

    # Iterate through the files and remove them
    for file_name in files_in_static:
        file_path = os.path.join(static_folder_path, file_name)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"Removed: {file_path}")
        except Exception as e:
            print(f"Error while removing {file_path}: {e}")

        print("All files in the 'static' folder have been removed.")
    else:
        print(
            "The 'static' folder does not exist in the current working directory."
        )
    try:
      remove_files_in_static_folder()
      
    except Exception as e:
      print(f"Error while removing {file_path}: {e}")
    try:
      shutil.rmtree(os.path.join(os.getcwd(), "static"))
    except Exception as e:
      print(f"Shutil Error while removing {file_path}: {e}")
